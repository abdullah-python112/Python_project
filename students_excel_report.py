from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

wb = Workbook()
sheet = wb.active
sheet.title = "Students"

students = [
    ("Abdullah", 21, "Latakia"),
    ("Ahmad", 18, "Damascus"),
    ("Batoul", 25, "Aleppo")
]
sheet["A1"] = "Name"
sheet["B1"] = "Age"
sheet["C1"] = "City"

for col in ["A1", "B1", "C1"]:
    sheet[col].font = Font(bold=True, size=14)
    sheet[col].fill = PatternFill(start_color="FFFF00", fill_type="solid")

for row, (name, age, city) in enumerate(students, start=2):
    sheet[f"A{row}"] = name
    sheet[f"B{row}"] = age
    sheet[f"C{row}"] = city

sheet.column_dimensions["A"].width = 15
sheet.column_dimensions["B"].width = 10
sheet.column_dimensions["C"].width = 20
wb.save("students.xlsx")

print(" تم إنشاء ملف Excel احترافي بنجاح!")