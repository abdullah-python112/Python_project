import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

with open("sales.csv", "r") as f:
    reader = csv.DictReader(f)
    data = list(reader)

wb = Workbook()
sheet = wb.active
sheet.title = "تقرير المبيعات"

headers = ["المنتج", "الكمية", "السعر"]
for col, header in enumerate(headers, start=1):
    cell = sheet.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True, size=14)
    cell.fill = PatternFill(start_color="FFFF00", fill_type="solid")

for i, row in enumerate(data, start=2):
    sheet.cell(row=i, column=1, value=row["Product"])
    sheet.cell(row=i, column=2, value=row["Qty"])
    sheet.cell(row=i, column=3, value=row["Price"])

wb.save("تقرير_المبيعات.xlsx")
print("✅ تم إنشاء التقرير بنجاح!")